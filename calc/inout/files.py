import os
import pickle

from pprint import pformat

import pandas as pd
import numpy as np

import openpyxl as oxl

from calc.setup import config as cfg
from calc.functions.checks import checkstr

# HARD-CODED
max_excel_rows = 800000 # actual limit: 1048576
max_excel_cols = 10000 # actual limit: 16384



def as_text(value):
    if value is None:
        return ""
    return str(value)

def add_full_directory(str):
    """
    if paths are relative, convert to path with drive letters
    :param str: path from which we start
    :return: full path with explicit drive letters
    """
    if str is None:
        return str
    else:
        if str[1] == ':':  # string does begin with drive letter and :\
            return str

        else:
            if str[0] == '\\':  # string does begin with a sep
                return os.path.realpath('.') + str
            else:
                return os.path.realpath('.') + os.sep + str


def make_filepathext(file, path='', ext=''):
    """
    given a file name, extension, and a path, return a full, explicit (with drive letter) path.
    if no path is given, the current directory for the .py files is used
    :param file:
    :param path:
    :param ext:
    :return:
    """
    # dealing with strings, which are immutable

    use_file = file
    use_path = path
    use_ext = ext

    # check the extension starts with '.'
    if len(use_ext) > 0:
        if use_ext[0] != '.':
            use_ext = '.' + use_ext

    # check the path ends with separator
    if len(use_path) > 0:
        use_path = add_full_directory(use_path)
        if use_path[-1] != os.sep:
            use_path = use_path + os.sep
    else:
        use_file = add_full_directory(file)

    #print(f'check_filepath output = {use_path + file + use_ext}')

    return use_path + use_file + use_ext

def get_filetablecol(datanameID):
    """
    Based on a nameID from data table, get the file name and file table (sheet or geo)
    Also return the column from the data table, since there's a bit of logic in
    figuring out if this is a geofile or regular file
    :param datanameID: a string that matches something in the index of t_data
    :return: tuple (name, table, col)
    """

    if datanameID in cfg.df_data.index:
        filenameID = cfg.df_data.loc[datanameID, cfg.tbl_data_col_sourceUniqueName]
        filetable = cfg.tbl_files
        col_with_data = cfg.tbl_data_col_headerData

        return filenameID, filetable, col_with_data

    elif datanameID in cfg.df_stms.index:
        stm_nameID = datanameID
        filetable = cfg.tbl_stms
        col_with_data = ''

        return stm_nameID, filetable, col_with_data

    else:
        return None, None, None

    # if datanameID in cfg.dict_excelTables[cfg.t_data].index:
    #
    #     # set empty variables
    #     filenameID = ''
    #     filetable = ''
    #     col_in_filetable = ''
    #
    #     # look in both tables
    #     # file name
    #     regname = cfg.dict_excelTables[cfg.t_data].loc[
    #         datanameID, cfg.s_data_regfilename]
    #     # geofile name
    #     geoname = cfg.dict_excelTables[cfg.t_data].loc[
    #         datanameID, cfg.s_data_geofilename]
    #
    #     # use the one that's not empty to get file name
    #     if regname is None:
    #         # not here
    #         pass
    #     elif len(regname) == 0:
    #         # not here
    #         pass
    #     else:
    #         # in the files tables, so we expect a column, or blank for a matrix
    #         filenameID = regname
    #         filetable = cfg.t_reg_files
    #         col_in_filetable = cfg.dict_excelTables[cfg.t_data].loc[datanameID, cfg.s_data_colregfile]
    #
    #     if geoname is None:
    #         # not here (nothing in the cell)
    #         pass
    #     elif len(geoname) == 0:
    #         # not here (empty string in the cell)
    #         pass
    #     else:
    #         # in the files tables
    #         filenameID = geoname
    #         filetable = cfg.t_geo_files
    #         col_in_filetable = cfg.dict_excelTables[cfg.t_data].loc[datanameID, cfg.s_data_colgeofile]
    #
    #     return filenameID, filetable, col_in_filetable
    #
    # else:
    #     raise KeyError(f'datanameID = "{datanameID}" was NOT in table Tbl_Data')


def get_filerow(datanameID):
    # given a datanameID, return the entire row from the corresponding table

    # get the row name and the table
    filenameID, filetable, _ = get_filetablecol(datanameID)

    return cfg.dict_excelTables[filetable].loc[filenameID, :]


def add_trailing_file_sep(strpath):
    if strpath[-1] != os.sep:
        strpath = strpath + os.sep
    return strpath

def combine_dir_subdir(main_path, sub_path):
    path = add_trailing_file_sep(main_path)
    if checkstr(sub_path):
        path = path + add_trailing_file_sep(sub_path)
    return path


def get_filepathext_from_files(filenameID):
    r"""
    Return tuple of directory/path information based on a filenameID
    in a given table (either the data or the gis files).
    For C:\directory\example.csv, these are the file name (just 'example'),
    the path ('C:\directory'), and extension ('.csv')
    :param filenameID: first column (the row index) in the data or gis tables
    :return: tuple of file, path, extension
    """
    if filenameID in cfg.df_files.index:
        datarow = cfg.df_files.loc[filenameID, :]

        # gotta check for trailing separators in the path and subfolder
        path = add_full_directory(datarow[cfg.tbl_files_col_fileDirMain])
        path = combine_dir_subdir(path, datarow[cfg.tbl_files_col_fileDirSub])

        # path = add_trailing_file_sep(path)
        #
        # folderpath = datarow[cfg.tbl_files_col_fileDirSub]
        #
        # # folder may be empty; if it's not, add and check the trailing separator
        # if not folderpath is None:
        #     if len(folderpath) > 0:
        #         path = path + folderpath
        #         path = add_trailing_file_sep(path)

        file_actual_name = datarow[cfg.tbl_files_col_fileName]
        ext = datarow[cfg.tbl_files_col_fileExt]

        return file_actual_name, path, ext
    else:
        raise KeyError(f'function <get_filepathext_from_files> could not find File_NameID = {filenameID} in the file table')


def get_filepathext_from_data(datanameID):
    # given a datanameID in t_data, get the file, path, and extension

    filenameID, table, _ = get_filetablecol(datanameID)

    fileshortname, path, ext = get_filepathext_from_files(filenameID)

    return fileshortname, path, ext


def get_write_mode(file, path='', ext=''):
    """
    check if a file exists, which will determine if we are writing a new file or adding to an existing
    :param file:
    :param path:
    :param ext:
    :return:
    """
    if os.path.exists(make_filepathext(file=file, path=path, ext=ext)):
        return 'a'  # append
    else:
        return 'w'  # write


def read_or_save_pickle(action, file, path='', ext='', list_save_vars=''):
    """
    a perhaps unnecessary file to help reading or saving pickles.
    :param action: text strings: 'save' or 'read', telling which to do
    :param file:
    :param path:
    :param ext:
    :param list_save_vars:
    :return:
    """
    fullpathname = make_filepathext(file=file, path=path, ext=ext)

    if action == 'save':
        with open(fullpathname, 'wb') as f:
            pickle.dump(list_save_vars, f)

    elif action == 'read':
        with open(fullpathname, 'rb') as f:
            unpickled = pickle.load(f)

        return unpickled

    else:
        # function not called correctly
        raise TypeError(f'Function <read_or_save_pickle> called without '
                        f'"read" or "save" parameter')


def save_df_or_array(data, path, filename, extension):
    """
    a multi-functional saving function, to save multiple types of files
    :param data:
    :param path:
    :param filename:
    :param extension:
    :return:
    """
    # save a dataframe or arrayt to a file

    bln_is_dataframe = isinstance(data, pd.DataFrame)  # was pd.core.frame.Dataframe
    bln_is_nparray = isinstance(data, np.ndarray)

    fullpathname = make_filepathext(file=filename, path=path, ext=extension)

    if not (bln_is_dataframe or bln_is_nparray):
        raise KeyError(
            f'Function <save_df_or_array> was called with an unknown data type.  '
            f'We can handle pandas dataframe or numpy array.  '
            f'This was Data type = {type(data)}')

    if extension == '.csv':
        if bln_is_dataframe:
            data.to_csv(path_or_buf=fullpathname,
                        sep=',')
        elif bln_is_nparray:
            np.savetxt(fname=fullpathname,
                       X=data, delimiter=',')
        else:
            # shouldn't be able to get here; we already checked for type
            pass

    elif extension == '.npy':
        if bln_is_dataframe:
            np.save(file=fullpathname, arr=data.to_numpy())
        elif bln_is_nparray:
            np.save(file=fullpathname, arr=data)
        else:
            # shouldn't be able to get here; we already checked for type
            pass

    elif extension == '.xlsx':
        if bln_is_dataframe:
            # if writing to existing file...
            # with pd.ExcelWriter(fullpathname, mode='w') as writer:
            #     data.to_excel(writer, sheet_name='pandas')
            data.to_excel(fullpathname,
                          sheet_name=filename, index=True, header=True)

        elif bln_is_nparray:
            data = pd.DataFrame(data)
            data.to_excel(excel_writer=fullpathname,
                          sheet_name=filename, index=True, header=True)
        else:
            # shouldn't be able to get here; we already checked for type
            pass
    else:
        raise KeyError(
            f'function <save_df_or_array> was called with an unknown extension.  '
            f'File = {filename + extension} in directory = {path}')



# region converting between dictionaries and text (mostly for checking contents)
def dict_to_text(dictionary):
    # pprint was a bit finicky... if width is too wide, it seemed that
    # some nested levels showed up on the same line
    return pformat(dictionary, width=180, compact=False)

def write_dictionary_to_txt(dictionary, fname):
    str_rep = dict_to_text(dictionary=dictionary)
    with open(fname + '.txt', "w") as text_file:
        text_file.write(str_rep)
# endregion


def write_df_to_excel(df, file, path='', ext='', sheet_name='', merge_cells=False):
    """
    a centralized file to nicely output dataframes to excel files.  Note the max excel rows and cols set above.
    :param df:
    :param file:
    :param path:
    :param ext:
    :param sheet_name:
    :param merge_cells:
    :return:
    """
    fullpathname = make_filepathext(file=file, path=path, ext=ext)
    writemode = get_write_mode(file=fullpathname)
    print (f'\t\t write_df_to_excel, with fullpath = {fullpathname}, writemode = {writemode}')

    if sheet_name == '':
        sheet_name = 'df_ffs'

    if df.shape[0] > max_excel_rows:
        bln_too_many_rows = True
        tempdf = df.iloc[0:100, :]
        msg = f'df_ffs size = {df.shape[0]} rows ... so not fully reported.  '
    else:
        bln_too_many_rows = False
        tempdf = df
        msg = ''

    if not isinstance(df, pd.Series):
        if df.shape[1] > max_excel_cols:
            bln_too_many_cols = True
            tempdf = tempdf.iloc[:,0:100]
            msg = msg + f'df_ffs size = {df.shape[1]} cols ... so not fully reported'
            bln_too_many_cols = True
        else:
            bln_too_many_cols = False
    else:
        bln_too_many_cols = False

    if bln_too_many_rows or bln_too_many_cols:
        tempdf.iloc[0,0] = msg

    #xlsxwriter may be faster for new spreadsheets

    if writemode == 'w':
        with pd.ExcelWriter(path=fullpathname, engine='xlsxwriter', mode=writemode) as writer:  #openpyxl engine needed to append
            tempdf.to_excel(writer, sheet_name=sheet_name, merge_cells=merge_cells)
    if writemode == 'a':
        with pd.ExcelWriter(path=fullpathname, engine='openpyxl', mode=writemode) as writer:  #openpyxl engine needed to append
            tempdf.to_excel(writer, sheet_name=sheet_name, merge_cells=merge_cells)


def write_dict_to_excel(dict, file, path='', ext='', sheet_name=''):
    str_rep = dict_to_text(dictionary=dict)

    write_text_to_excel(thetext=str_rep, file_name=file, path=path,
                        ext=ext, sheet_name=sheet_name)


def write_text_to_excel(thetext, file_name, path='', ext='', sheet_name=''):
    # dump it all into one cell

    fullpathname = make_filepathext(file=file_name, path=path, ext=ext)
    writemode = get_write_mode(file=fullpathname)

    #print(f'write_text_to_excel fullpathname={fullpathname}')

    if writemode == 'a':
        wb = oxl.load_workbook(filename=fullpathname)
        if sheet_name != '':
            if sheet_name in wb.worksheets:
                print('duplicate sheet... openpyxl will rename')
    else:
        wb = oxl.Workbook()

        if sheet_name == '':
            sheet_name = 'Sheet'

    ws = wb.create_sheet(title=sheet_name)

    if thetext.count('\n') > max_excel_rows:
        stop_row = 12
    else:
        stop_row = max_excel_rows

    row_count = 1
    for line in thetext.splitlines():
        ws.cell(row_count, column=1).value = line
        row_count += 1
        if row_count > stop_row:
            ws.cell(row_count, column=1).value = f'too many lines, so not reported'
            break

    # ws['A1'].value = thetext
    # ws['A1'].alignment = oxl.styles.Alignment(wrap_text=True)
    # ws.column_dimensions['A'].width = 80

    # https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
    # for column_cells in ws.columns:
    #     length = max(len(as_text(cell.value)) for cell in column_cells)
    #     ws.column_dimensions[column_cells[0].column].width = length

    wb.save(filename=fullpathname)



